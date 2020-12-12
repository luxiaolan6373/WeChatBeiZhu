import openpyxl,os
class KeHuGuanLi():
    def __init__(self,filename):
        self.filename = filename
        try:#判断是否有这个xml 没有就创建一个
            self.wb = openpyxl.load_workbook(filename)

        except:
            # 判断是否有这个目录,没有的话就建立
            if os.path.isdir(path) != True:
                os.makedirs('datas')
                self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        # 指定一个单元格,它上方和坐标的单元格将全部定住
        self.ws.freeze_panes = 'B2'

    def save(self):
        self.wb.save(self.filename)

    def getLabel(self):
        '''
        获取所有标签的文本
        :return: 会返回一个列表
        '''
        ls=[]
        #print(self.ws.max_column)
        for item in self.ws.columns:
            ls.append(item[0].value)
        return ls
    def getRow(self):
        '''
        取最大行
        :return:
        '''

        return self.ws.max_row

    def getColumn(self):
        '''
        取最大列
        :return:
        '''
        return  self.ws.max_column


if __name__ == '__main__':
    khxlsx=KeHuGuanLi()





